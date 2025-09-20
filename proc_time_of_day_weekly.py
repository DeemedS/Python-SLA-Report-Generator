from openpyxl.worksheet.worksheet import Worksheet
from conn import get_connection
import logging
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from datetime import datetime, timedelta

# Define styles
title_font = Font(bold=True, color="008000", size=11)  # Green bold title
header_font = Font(bold=True, color="FFFFFF")          # White bold header
header_fill = PatternFill(start_color="156082", end_color="156082", fill_type="solid")  # Light green fill
border = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin")
)

def generate_proc_time_of_day_weekly(
    sheet: Worksheet,
    end_date: str, 
    logger: logging.Logger
) -> bool:
    """Generate TIME OF DAY - TXN_COUNT (DAILY) data into the given Excel sheet."""
    logger.info("Generating TIME OF DAY - TXN_COUNT (WEEKLY) data...")

    # Parse end_date
    end_dt = datetime.strptime(end_date, "%Y-%m-%d")

    # Get the first day of the month
    start_dt = end_dt.replace(day=1)

    # Find the first Sunday on or before the start date
    start_week_dt = start_dt - timedelta(days=start_dt.weekday() + 1 if start_dt.weekday() != 6 else 0)
    start_week = start_week_dt.strftime("%Y-%m-%d")

    # MIN AND MAX DATES
    min_date = start_week
    max_date = end_date

    # MIN WEEK AND MAX WEEK FROM DB
    min_week_query = """
        SELECT MIN(WEEK_NUM) 
        FROM [dbo].[txn_analysis_time_of_day_txn_count]
        WHERE TRANSACTION_DATE BETWEEN ? AND ?
    """

    max_week_query = """
        SELECT MAX(WEEK_NUM) 
        FROM [dbo].[txn_analysis_time_of_day_txn_count]
        WHERE TRANSACTION_DATE BETWEEN ? AND ?
    """
    try:
        conn = get_connection()
        cursor = conn.cursor()
        cursor.execute(min_week_query, (min_date, max_date))
        min_week = cursor.fetchone()[0]

        cursor.execute(max_week_query, (min_date, max_date))
        max_week = cursor.fetchone()[0]
        cursor.close()
        conn.close()
    except Exception as e:
        logger.error(f"Error min an max week for: {e}")
        return False
    
    logger.info(f"Min week: {min_week}, Max week: {max_week}")
    
    # Add title
    title_cell = sheet.cell(row=1, column=2, value="TRANSACTION COUNT PER TIME OF DAY AND FINAL STATUS")
    title_cell.font = title_font

    for week_num in range(min_week, max_week + 1):
        logger.info(f"Processing WEEK_NUM: {week_num}")

        min_date_week = """
            SELECT MIN(TRANSACTION_DATE) 
            FROM [dbo].[txn_analysis_time_of_day_txn_count]
            WHERE TRANSACTION_DATE BETWEEN ? AND ?
            AND WEEK_NUM = ?
        """

        max_date_week = """
            SELECT MAX(TRANSACTION_DATE) 
            FROM [dbo].[txn_analysis_time_of_day_txn_count]
            WHERE TRANSACTION_DATE BETWEEN ? AND ?
            AND WEEK_NUM = ?
        """

        try:
            conn = get_connection()
            cursor = conn.cursor()
            cursor.execute(min_date_week, (min_date, max_date, week_num))
            min_date_of_week = cursor.fetchone()[0]

            cursor.execute(max_date_week, (min_date, max_date, week_num))
            max_date_of_week = cursor.fetchone()[0]
            cursor.close()
            conn.close()
        except Exception as e:
            logger.error(f"Error min an max week for: {e}")
            return False
        
        logger.info(f"Min date: {min_date_of_week}, Max date : {max_date_of_week}")

        # First Block 
        try:
            conn = get_connection()
            cursor = conn.cursor()

            logger.info(f"Fetching data for first block week {week_num} from database for first block...")

            cursor.execute("""
                SELECT 
                [WEEK_NUM],
                [FINAL_STATUS],
                SUM([0-8]) AS [0-8],
                SUM([9-12]) AS [9-12],
                SUM([13-16]) AS [13-16],
                SUM([17-20]) AS [17-20],
                SUM([21-23]) AS [21-23],
                SUM([0-8]) + SUM([9-12]) + SUM([13-16]) + SUM([17-20])  + SUM([21-23])
                FROM [dbo].[txn_analysis_time_of_day_txn_count]
                WHERE TRANSACTION_DATE BETWEEN ? AND ?
                AND WEEK_NUM = ?
                GROUP BY WEEK_NUM, FINAL_STATUS
                ORDER BY WEEK_NUM, FINAL_STATUS
            """, min_date_of_week, max_date_of_week, week_num)

            rows = cursor.fetchall()

            # Add headers 
            columns = [col[0] for col in cursor.description]
            columns.insert(0, f"{min_date_of_week} - {max_date_of_week}")
            sheet.append(columns)

            # Style header row
            header_row_idx = sheet.max_row
            for col_idx, _ in enumerate(columns, start=1):
                cell = sheet.cell(row=header_row_idx, column=col_idx)
                if col_idx > 1 and col_idx < 9:
                    cell.font = header_font
                    cell.fill = header_fill
                    cell.border = border
                    cell.alignment = Alignment(horizontal="center")

            # Append each row
            for row in rows:
                sheet.append(('',) + tuple(row))
                row_idx = sheet.max_row
                for col_idx in range(1, len(columns) + 1):
                    cell = sheet.cell(row=row_idx, column=col_idx)
                    if col_idx > 1 and col_idx < 9:
                        cell.alignment = Alignment(horizontal="center")
                        cell.border = border

            summary_query_vertical = """
                SELECT 
                '',
                '',
                SUM([0-8]) AS [0-8],
                SUM([9-12]) AS [9-12],
                SUM([13-16]) AS [13-16],
                SUM([17-20]) AS [17-20],
                SUM([21-23]) AS [21-23]
                FROM [dbo].[txn_analysis_time_of_day_txn_count]
                WHERE TRANSACTION_DATE BETWEEN ? AND ?
                AND WEEK_NUM = ?
            """

            # Execute Vertical Summary
            cursor.execute(summary_query_vertical, min_date_of_week, max_date_of_week, week_num)
            summary_row_vertical = cursor.fetchone()

            # Append vertaical summary
            if summary_row_vertical:
                row_idx = sheet.max_row + 1
                for col_idx, value in enumerate(summary_row_vertical, start=2):
                    sheet.cell(row=row_idx, column=col_idx, value=value)

            cursor.close()
            conn.close()
            
        except Exception as e:
            logger.error(f"Error fetching data for week {week_num}: {e}")
            return False

        #Second Block
        start_col_second = 11

        try:
            conn = get_connection()
            cursor = conn.cursor()
            logger.info(f"Fetching data for for week {week_num} from database...")
            
            cursor.execute("""
            SELECT 
                FINAL_STATUS,
                CAST(SUM([0-8]) * 1.0 / NULLIF(SUM(SUM([0-8])) OVER (), 0) AS DECIMAL(10,6)) AS [0-8],
                CAST(SUM([9-12]) * 1.0 / NULLIF(SUM(SUM([9-12])) OVER (), 0) AS DECIMAL(10,6)) AS [9-12],
                CAST(SUM([13-16]) * 1.0 / NULLIF(SUM(SUM([13-16])) OVER (), 0) AS DECIMAL(10,6)) AS [13-16],
                CAST(SUM([17-20]) * 1.0 / NULLIF(SUM(SUM([17-20])) OVER (), 0) AS DECIMAL(10,6)) AS [17-20],
                CAST(SUM([21-23]) * 1.0 / NULLIF(SUM(SUM([21-23])) OVER (), 0) AS DECIMAL(10,6)) AS [21-23]
            FROM [dbo].[txn_analysis_time_of_day_txn_count]
            WHERE TRANSACTION_DATE BETWEEN ? AND ?
            AND WEEK_NUM = ?
            GROUP BY FINAL_STATUS
            ORDER BY FINAL_STATUS;
            """, min_date_of_week, max_date_of_week, week_num)

            rows = cursor.fetchall()

            summary_query_vertical = """
                WITH totals AS (
                    SELECT
                        SUM([0-8])  AS sum_0_8,
                        SUM([9-12]) AS sum_9_12,
                        SUM([13-16]) AS sum_13_16,
                        SUM([17-20]) AS sum_17_20,
                        SUM([21-23]) AS sum_21_23
                    FROM [dbo].[txn_analysis_time_of_day_txn_count]
                    WHERE TRANSACTION_DATE BETWEEN ? AND ?
                    AND WEEK_NUM = ?
                )
                SELECT
                    '',
                    CAST(sum_0_8  * 1.0 / NULLIF(sum_0_8, 0)   AS DECIMAL(10,6)) AS [0-8],
                    CAST(sum_9_12 * 1.0 / NULLIF(sum_9_12, 0)  AS DECIMAL(10,6)) AS [9-12],
                    CAST(sum_13_16 * 1.0 / NULLIF(sum_13_16, 0) AS DECIMAL(10,6)) AS [13-16],
                    CAST(sum_17_20 * 1.0 / NULLIF(sum_17_20, 0) AS DECIMAL(10,6)) AS [17-20],
                    CAST(sum_21_23 * 1.0 / NULLIF(sum_21_23, 0) AS DECIMAL(10,6)) AS [21-23]
                FROM totals;
            """

            # Add headers 
            columns = [col[0] for col in cursor.description]

            # Style header row
            header_row_idx = sheet.max_row - len(rows) - 1
            for col_idx, header in enumerate(columns, start=start_col_second):
                cell = sheet.cell(row=header_row_idx, column=col_idx, value=header)
                if col_idx > 10 and col_idx < 18:
                    cell.font = header_font
                    cell.fill = header_fill
                    cell.border = border
                    cell.alignment = Alignment(horizontal="center")

            for i, row in enumerate(rows, start=1):
                row_idx = header_row_idx + i
                for col_idx, value in enumerate(tuple(row), start=start_col_second):
                    cell = sheet.cell(row=row_idx, column=col_idx, value=value)
                    if col_idx >= start_col_second and col_idx < start_col_second + 6:
                        cell.alignment = Alignment(horizontal="center")
                        cell.border = border
                        
                    if 11 <= col_idx <= 16:
                        cell.number_format = "0.00%"

            # Execute Vertical Summary
            cursor.execute(summary_query_vertical, min_date_of_week, max_date_of_week, week_num)
            summary_row_vertical = cursor.fetchone()

            
            if summary_row_vertical:
                row_idx = sheet.max_row
                for col_idx, value in enumerate(summary_row_vertical, start=start_col_second):
                    cell = sheet.cell(row=row_idx, column=col_idx, value=value)
                    cell.number_format = '0.00%'

            cursor.close()
            conn.close()
        except Exception as e:
            logger.error(f"Error fetching data for week {week_num}: {e}")
            return False

        #Third Block
        start_col_third = 18

        try:
            conn = get_connection()
            cursor = conn.cursor()
            logger.info(f"Fetching data for week {week_num} from database...")
            cursor.execute("""
                SELECT
                    FINAL_STATUS,
                    CAST(SUM([0-8])   * 1.0 / NULLIF((SUM([0-8])   + SUM([9-12]) + SUM([13-16]) + SUM([17-20]) + SUM([21-23])), 0) AS DECIMAL(10,6)) AS [0-8],
                    CAST(SUM([9-12])  * 1.0 / NULLIF((SUM([0-8])   + SUM([9-12]) + SUM([13-16]) + SUM([17-20]) + SUM([21-23])), 0) AS DECIMAL(10,6)) AS [9-12],
                    CAST(SUM([13-16]) * 1.0 / NULLIF((SUM([0-8])   + SUM([9-12]) + SUM([13-16]) + SUM([17-20]) + SUM([21-23])), 0) AS DECIMAL(10,6)) AS [13-16],
                    CAST(SUM([17-20]) * 1.0 / NULLIF((SUM([0-8])   + SUM([9-12]) + SUM([13-16]) + SUM([17-20]) + SUM([21-23])), 0) AS DECIMAL(10,6)) AS [17-20],
                    CAST(SUM([21-23]) * 1.0 / NULLIF((SUM([0-8])   + SUM([9-12]) + SUM([13-16]) + SUM([17-20]) + SUM([21-23])), 0) AS DECIMAL(10,6)) AS [21-23],
                    CAST(1.0 AS DECIMAL(10,6))
                FROM [dbo].[txn_analysis_time_of_day_txn_count]
                WHERE TRANSACTION_DATE BETWEEN ? AND ?
                AND WEEK_NUM = ?
                GROUP BY FINAL_STATUS
                ORDER BY FINAL_STATUS
            """, min_date_of_week, max_date_of_week, week_num)
            rows = cursor.fetchall()

            # Add headers 
            columns = [col[0] for col in cursor.description]

            # Style header row
            header_row_idx = sheet.max_row - len(rows) - 1
            for col_idx, header in enumerate(columns, start=start_col_third):
                cell = sheet.cell(row=header_row_idx, column=col_idx, value=header)
                if col_idx > 17 and col_idx < 24:
                    cell.font = header_font
                    cell.fill = header_fill
                    cell.border = border
                    cell.alignment = Alignment(horizontal="center")

            for i, row in enumerate(rows, start=1):
                row_idx = header_row_idx + i
                for col_idx, value in enumerate(tuple(row), start=start_col_third):
                    cell = sheet.cell(row=row_idx, column=col_idx, value=value)
                    if col_idx >= start_col_third and col_idx < start_col_third + 6:
                        cell.alignment = Alignment(horizontal="center")
                        cell.border = border

                    if 18 <= col_idx <= 24:
                        cell.number_format = "0.00%"

            cursor.close()
            conn.close()
        except Exception as e:
            logger.error(f"Error fetching data for week {week_num}: {e}")
            return False
        
        # Add a blank row after each day's block
        if rows:
            sheet.append([])
        
    logger.info("TIME OF DAY - TXN_COUNT (WEELY) data generation complete.")
    return True

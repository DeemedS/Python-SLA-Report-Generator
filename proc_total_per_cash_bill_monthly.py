from openpyxl.worksheet.worksheet import Worksheet
from conn import get_connection
import logging
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment

# Define styles
title_font = Font(bold=True, color="008000", size=11)  # Green bold title
header_font = Font(bold=True, color="FFFFFF")          # White bold header
header_fill = PatternFill(start_color="156082", end_color="156082", fill_type="solid")  # Light green fill
border = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin")
)

def generate_total_per_cash_bill_monthly(
    sheet: Worksheet,
    end_date: str, 
    logger: logging.Logger,
    year_start: str
) -> bool:
    """Generate TOTAL_PER CASH BILL - CBI (MONTHLY) data into the given Excel sheet."""
    logger.info("Generating TOTAL_PER CASH BILL - CBI (MONTHLY) data...")

    # MIN AND MAX DATES
    min_date = year_start
    max_date = end_date

    # MIN WEEK AND MAX WEEK FROM DB
    min_month_query = """
        SELECT MIN([MONTH]) 
        FROM [dbo].[txn_analysis_total_per_cash_bill_cbi_txn]
        WHERE TRANSACTION_DATE BETWEEN ? AND ?
    """

    max_month_query = """
        SELECT MAX([MONTH]) 
        FROM [dbo].[txn_analysis_total_per_cash_bill_cbi_txn]
        WHERE TRANSACTION_DATE BETWEEN ? AND ?
    """
    try:
        conn = get_connection()
        cursor = conn.cursor()
        cursor.execute(min_month_query, (min_date, max_date))
        min_month = cursor.fetchone()[0]

        cursor.execute(max_month_query, (min_date, max_date))
        max_month = cursor.fetchone()[0]
        cursor.close()
        conn.close()
    except Exception as e:
        logger.error(f"Error min an max week for: {e}")
        return False
    
    logger.info(f"Min month: {min_month}, Max month: {max_month}")

    # TXN LOOP 
    logger.info("Processing COUNT")
    
    # Add title
    title_cell = sheet.cell(row=1, column=2, value="COUNT PER BILL INSERTED AND FINAL STATUS")
    title_cell.font = title_font

    for month in range(min_month, max_month + 1):
        logger.info(f"Processing MONTH: {month}")

        min_date_month_query = """
            SELECT MIN(TRANSACTION_DATE) 
            FROM [dbo].[txn_analysis_total_per_cash_bill_cbi_txn]
            WHERE TRANSACTION_DATE BETWEEN ? AND ?
            AND [MONTH] = ?
        """

        max_date_month_query = """
            SELECT MAX(TRANSACTION_DATE) 
            FROM [dbo].[txn_analysis_total_per_cash_bill_cbi_txn]
            WHERE TRANSACTION_DATE BETWEEN ? AND ?
            AND [MONTH] = ?
        """

        try:
            conn = get_connection()
            cursor = conn.cursor()
            cursor.execute(min_date_month_query, (min_date, max_date, month))
            min_date_month = cursor.fetchone()[0]

            cursor.execute(max_date_month_query, (min_date, max_date, month))
            max_date_month = cursor.fetchone()[0]
            cursor.close()
            conn.close()
        except Exception as e:
            logger.error(f"Error min an max week for: {e}")
            return False
        
        logger.info(f"Min date: {min_date_month}, Max date : {max_date_month}")

        # First Block
        try:
            conn = get_connection()
            cursor = conn.cursor()
            logger.info(f"Fetching data for first block month {month} from database for first block...")
            cursor.execute("""
                SELECT 
                [MONTH],
                [FINAL_STATUS],
                SUM([P20_DENOM]) AS [P20_DENOM],
                SUM([P50_DENOM]) AS [P50_DENOM],
                SUM([P100_DENOM]) AS [P100_DENOM],
                SUM([P200_DENOM]) AS [P200_DENOM],
                SUM([P500_DENOM]) AS [P500_DENOM],
                SUM([P1000_DENOM]) AS [P1000_DENOM],
                SUM([P20_DENOM] + [P50_DENOM]+ [P100_DENOM] + [P200_DENOM] + [P500_DENOM] + [P1000_DENOM])
                FROM [dbo].[txn_analysis_total_per_cash_bill_cbi_txn]
                WHERE TRANSACTION_DATE BETWEEN ? AND ?
                AND [MONTH] = ?
                GROUP BY [MONTH], FINAL_STATUS
                ORDER BY [MONTH], FINAL_STATUS
            """, min_date_month, max_date_month, month)
            rows = cursor.fetchall()

            summary_query_vertical = """
                SELECT 
                '',
                '',
                SUM([P20_DENOM]) AS [P20_DENOM],
                SUM([P50_DENOM]) AS [P50_DENOM],
                SUM([P100_DENOM]) AS [P100_DENOM],
                SUM([P200_DENOM]) AS [P200_DENOM],
                SUM([P500_DENOM]) AS [P500_DENOM],
                SUM([P1000_DENOM]) AS [P1000_DENOM]
                FROM [dbo].[txn_analysis_total_per_cash_bill_cbi_txn]
                WHERE TRANSACTION_DATE BETWEEN ? AND ?
                AND [MONTH] = ?
            """

            # Add headers 
            columns = [col[0] for col in cursor.description]
            columns.insert(0, f"{min_date_month} - {max_date_month}")
            sheet.append(columns)

            # Style header row
            header_row_idx = sheet.max_row
            for col_idx, _ in enumerate(columns, start=1):
                cell = sheet.cell(row=header_row_idx, column=col_idx)
                if col_idx > 1 and col_idx < 10:
                    cell.font = header_font
                    cell.fill = header_fill
                    cell.border = border
                    cell.alignment = Alignment(horizontal="center")

            # Append each row
            for row in rows:
                sheet.append(('',) + tuple(row))
                row_idx = sheet.max_row
                for col_idx in range(1, len(columns) + 1):
                    cell = sheet.cell(row=row_idx, column=col_idx)
                    if col_idx > 1 and col_idx < 10:
                        cell.alignment = Alignment(horizontal="center")
                        cell.border = border

            # Execute Vertical Summary
            cursor.execute(summary_query_vertical, min_date_month, max_date_month, month)
            summary_row_vertical = cursor.fetchone()

            # Append vertaical summary
            if summary_row_vertical:
                row_idx = sheet.max_row + 1
                for col_idx, value in enumerate(summary_row_vertical, start=2):
                    sheet.cell(row=row_idx, column=col_idx, value=value)

            cursor.close()
            conn.close()
        except Exception as e:
            logger.error(f"Error fetching data for month {month}: {e}")
            return False
        
        # Second Block
        start_col_second = 12

        try:
            conn = get_connection()
            cursor = conn.cursor()
            logger.info(f"Fetching data for for month {month} from database...")
            
            cursor.execute("""
                SELECT 
                    FINAL_STATUS,
                    CAST(SUM([P20_DENOM]) * 1.0 / NULLIF(SUM(SUM([P20_DENOM])) OVER (), 0) AS DECIMAL(10,6)) AS [P20_DENOM],
                    CAST(SUM([P50_DENOM]) * 1.0 / NULLIF(SUM(SUM([P50_DENOM])) OVER (), 0) AS DECIMAL(10,6)) AS [P50_DENOM],
                    CAST(SUM([P100_DENOM]) * 1.0 / NULLIF(SUM(SUM([P100_DENOM])) OVER (), 0) AS DECIMAL(10,6)) AS [P100_DENOM],
                    CAST(SUM([P200_DENOM]) * 1.0 / NULLIF(SUM(SUM([P200_DENOM])) OVER (), 0) AS DECIMAL(10,6)) AS [P200_DENOM],
                    CAST(SUM([P500_DENOM]) * 1.0 / NULLIF(SUM(SUM([P500_DENOM])) OVER (), 0) AS DECIMAL(10,6)) AS [P500_DENOM],
                    CAST(SUM([P1000_DENOM]) * 1.0 / NULLIF(SUM(SUM([P1000_DENOM])) OVER (), 0) AS DECIMAL(10,6)) AS [P1000_DENOM]            
                FROM [dbo].[txn_analysis_total_per_cash_bill_cbi_txn]
                WHERE TRANSACTION_DATE BETWEEN ? AND ?
                AND [MONTH] = ?
                GROUP BY FINAL_STATUS
                ORDER BY FINAL_STATUS;
            """, min_date_month, max_date_month, month)
            rows = cursor.fetchall()

            summary_query_vertical = """
                WITH totals AS (
                    SELECT
                        SUM([P20_DENOM]) AS [SUM_P20_DENOM],
                        SUM([P50_DENOM]) AS [SUM_P50_DENOM],
                        SUM([P100_DENOM]) AS [SUM_P100_DENOM],
                        SUM([P200_DENOM]) AS [SUM_P200_DENOM],
                        SUM([P500_DENOM]) AS [SUM_P500_DENOM],
                        SUM([P1000_DENOM]) AS [SUM_P1000_DENOM] 
                    FROM [dbo].[txn_analysis_total_per_cash_bill_cbi_txn]
                    WHERE TRANSACTION_DATE BETWEEN ? AND ?
                    AND [MONTH] = ?
                )
                SELECT
                    '',
                    CAST([SUM_P20_DENOM]  * 1.0 / NULLIF([SUM_P20_DENOM], 0)   AS DECIMAL(10,6)) AS [P20_DENOM],
                    CAST([SUM_P50_DENOM] * 1.0 / NULLIF([SUM_P50_DENOM], 0)  AS DECIMAL(10,6)) AS [P50_DENOM],
                    CAST([SUM_P100_DENOM] * 1.0 / NULLIF([SUM_P100_DENOM], 0) AS DECIMAL(10,6)) AS [P100_DENOM],
                    CAST([SUM_P200_DENOM] * 1.0 / NULLIF([SUM_P200_DENOM], 0) AS DECIMAL(10,6)) AS [P200_DENOM],
                    CAST([SUM_P500_DENOM] * 1.0 / NULLIF([SUM_P500_DENOM], 0) AS DECIMAL(10,6)) AS [P500_DENOM],
                    CAST([SUM_P1000_DENOM] * 1.0 / NULLIF([SUM_P1000_DENOM], 0) AS DECIMAL(10,6)) AS [P1000_DENOM]
                FROM totals;
            """

            # Add headers 
            columns = [col[0] for col in cursor.description]

            # Style header row
            header_row_idx = sheet.max_row - len(rows) - 1
            for col_idx, header in enumerate(columns, start=start_col_second):
                cell = sheet.cell(row=header_row_idx, column=col_idx, value=header)
                if col_idx > 11  and col_idx < 20:
                    cell.font = header_font
                    cell.fill = header_fill
                    cell.border = border
                    cell.alignment = Alignment(horizontal="center")

            for i, row in enumerate(rows, start=1):
                row_idx = header_row_idx + i
                for col_idx, value in enumerate(tuple(row), start=start_col_second):
                    cell = sheet.cell(row=row_idx, column=col_idx, value=value)
                    if col_idx >= start_col_second and col_idx < start_col_second + 7:
                        cell.alignment = Alignment(horizontal="center")
                        cell.border = border
                        
                    if 13 <= col_idx <= 18:
                        cell.number_format = "0.00%"

            # Execute Vertical Summary
            cursor.execute(summary_query_vertical, min_date_month, max_date_month, month)
            summary_row_vertical = cursor.fetchone()
            
            if summary_row_vertical:
                row_idx = sheet.max_row
                for col_idx, value in enumerate(summary_row_vertical, start=start_col_second):
                    cell = sheet.cell(row=row_idx, column=col_idx, value=value)
                    cell.number_format = '0.00%'

            cursor.close()
            conn.close()
        except Exception as e:
            logger.error(f"Error fetching data for month {month}: {e}")
            return False
        
        #Third Block
        start_col_third = 20

        try:
            conn = get_connection()
            cursor = conn.cursor()
            logger.info(f"Fetching data for month {month} from database...")
            cursor.execute("""
                SELECT
                    FINAL_STATUS,
                    CAST(SUM([P20_DENOM])     * 1.0 / NULLIF((SUM([P20_DENOM]) + SUM([P50_DENOM]) + SUM([P100_DENOM]) + SUM([P200_DENOM]) + SUM([P500_DENOM]) + SUM([P1000_DENOM])), 0) AS DECIMAL(10,6)) AS [P20_DENOM],
                    CAST(SUM([P50_DENOM])     * 1.0 / NULLIF((SUM([P20_DENOM]) + SUM([P50_DENOM]) + SUM([P100_DENOM]) + SUM([P200_DENOM]) + SUM([P500_DENOM]) + SUM([P1000_DENOM])), 0) AS DECIMAL(10,6)) AS [P50_DENOM],
                    CAST(SUM([P100_DENOM])    * 1.0 / NULLIF((SUM([P20_DENOM]) + SUM([P50_DENOM]) + SUM([P100_DENOM]) + SUM([P200_DENOM]) + SUM([P500_DENOM]) + SUM([P1000_DENOM])), 0) AS DECIMAL(10,6)) AS [P100_DENOM],
                    CAST(SUM([P200_DENOM])    * 1.0 / NULLIF((SUM([P20_DENOM]) + SUM([P50_DENOM]) + SUM([P100_DENOM]) + SUM([P200_DENOM]) + SUM([P500_DENOM]) + SUM([P1000_DENOM])), 0) AS DECIMAL(10,6)) AS [P200_DENOM],
                    CAST(SUM([P500_DENOM])    * 1.0 / NULLIF((SUM([P20_DENOM]) + SUM([P50_DENOM]) + SUM([P100_DENOM]) + SUM([P200_DENOM]) + SUM([P500_DENOM]) + SUM([P1000_DENOM])), 0) AS DECIMAL(10,6)) AS [P500_DENOM],
                    CAST(SUM([P1000_DENOM])   * 1.0 / NULLIF((SUM([P20_DENOM]) + SUM([P50_DENOM]) + SUM([P100_DENOM]) + SUM([P200_DENOM]) + SUM([P500_DENOM]) + SUM([P1000_DENOM])), 0) AS DECIMAL(10,6)) AS [P1000_DENOM],
                    CAST(1.0 AS DECIMAL(10,6))
                FROM [dbo].[txn_analysis_total_per_cash_bill_cbi_txn]
                WHERE TRANSACTION_DATE BETWEEN ? AND ?
                AND [MONTH] = ?
                GROUP BY FINAL_STATUS
                ORDER BY FINAL_STATUS
            """, min_date_month, max_date_month, month)
            rows = cursor.fetchall()

            # Add headers 
            columns = [col[0] for col in cursor.description]

            # Style header row
            header_row_idx = sheet.max_row - len(rows) - 1
            for col_idx, header in enumerate(columns, start=start_col_third):
                cell = sheet.cell(row=header_row_idx, column=col_idx, value=header)
                if col_idx > 15 and col_idx < 27:
                    cell.font = header_font
                    cell.fill = header_fill
                    cell.border = border
                    cell.alignment = Alignment(horizontal="center")

            for i, row in enumerate(rows, start=1):
                row_idx = header_row_idx + i
                for col_idx, value in enumerate(tuple(row), start=start_col_third):
                    cell = sheet.cell(row=row_idx, column=col_idx, value=value)
                    if col_idx >= start_col_third and col_idx < start_col_third + 7:
                        cell.alignment = Alignment(horizontal="center")
                        cell.border = border

                    if 16 <= col_idx <= 27:
                        cell.number_format = "0.00%"

            cursor.close()
            conn.close()
        except Exception as e:
            logger.error(f"Error fetching data for month {month}: {e}")
            return False

        
        # Add a blank row after each day's block
        if rows:
            sheet.append([])

    # VOLUME LOOP

    logger.info("Processing Volume")

    last_row = sheet.max_row + 2  
    new_title_cell = sheet.cell(row=last_row, column=2, value="TOTAL VOLUME PER BILL INSERTED AND FINAL STATUS")
    new_title_cell.font = title_font
    
    sheet.append([])

    for month in range(min_month, max_month + 1):
        logger.info(f"Processing MONTH: {month}")

        min_date_month_query = """
            SELECT MIN(TRANSACTION_DATE) 
            FROM [dbo].[txn_analysis_total_per_cash_bill_cbi_vol]
            WHERE TRANSACTION_DATE BETWEEN ? AND ?
            AND [MONTH] = ?
        """

        max_date_month_query = """
            SELECT MAX(TRANSACTION_DATE) 
            FROM [dbo].[txn_analysis_total_per_cash_bill_cbi_vol]
            WHERE TRANSACTION_DATE BETWEEN ? AND ?
            AND [MONTH] = ?
        """

        try:
            conn = get_connection()
            cursor = conn.cursor()
            cursor.execute(min_date_month_query, (min_date, max_date, month))
            min_date_month = cursor.fetchone()[0]

            cursor.execute(max_date_month_query, (min_date, max_date, month))
            max_date_month = cursor.fetchone()[0]
            cursor.close()
            conn.close()
        except Exception as e:
            logger.error(f"Error min an max week for: {e}")
            return False
        
        logger.info(f"Min date: {min_date_month}, Max date : {max_date_month}")

        # First Block
        try:
            conn = get_connection()
            cursor = conn.cursor()
            logger.info(f"Fetching data for first block month {month} from database for first block...")
            cursor.execute("""
                SELECT 
                [MONTH],
                [FINAL_STATUS],
                SUM([P20_DENOM]) AS [P20_DENOM],
                SUM([P50_DENOM]) AS [P50_DENOM],
                SUM([P100_DENOM]) AS [P100_DENOM],
                SUM([P200_DENOM]) AS [P200_DENOM],
                SUM([P500_DENOM]) AS [P500_DENOM],
                SUM([P1000_DENOM]) AS [P1000_DENOM],
                SUM([P20_DENOM] + [P50_DENOM]+ [P100_DENOM] + [P200_DENOM] + [P500_DENOM] + [P1000_DENOM])
                FROM [dbo].[txn_analysis_total_per_cash_bill_cbi_vol]
                WHERE TRANSACTION_DATE BETWEEN ? AND ?
                AND [MONTH] = ?
                GROUP BY [MONTH], FINAL_STATUS
                ORDER BY [MONTH], FINAL_STATUS
            """, min_date_month, max_date_month, month)
            rows = cursor.fetchall()

            summary_query_vertical = """
                SELECT 
                '',
                '',
                SUM([P20_DENOM]) AS [P20_DENOM],
                SUM([P50_DENOM]) AS [P50_DENOM],
                SUM([P100_DENOM]) AS [P100_DENOM],
                SUM([P200_DENOM]) AS [P200_DENOM],
                SUM([P500_DENOM]) AS [P500_DENOM],
                SUM([P1000_DENOM]) AS [P1000_DENOM]
                FROM [dbo].[txn_analysis_total_per_cash_bill_cbi_vol]
                WHERE TRANSACTION_DATE BETWEEN ? AND ?
                AND [MONTH] = ?
            """

            # Add headers 
            columns = [col[0] for col in cursor.description]
            columns.insert(0, f"{min_date_month} - {max_date_month}")
            sheet.append(columns)

            # Style header row
            header_row_idx = sheet.max_row
            for col_idx, _ in enumerate(columns, start=1):
                cell = sheet.cell(row=header_row_idx, column=col_idx)
                if col_idx > 1 and col_idx < 10:
                    cell.font = header_font
                    cell.fill = header_fill
                    cell.border = border
                    cell.alignment = Alignment(horizontal="center")

            # Append each row
            for row in rows:
                sheet.append(('',) + tuple(row))
                row_idx = sheet.max_row
                for col_idx in range(1, len(columns) + 1):
                    cell = sheet.cell(row=row_idx, column=col_idx)
                    if col_idx > 1 and col_idx < 10:
                        cell.alignment = Alignment(horizontal="center")
                        cell.border = border

            # Execute Vertical Summary
            cursor.execute(summary_query_vertical, min_date_month, max_date_month, month)
            summary_row_vertical = cursor.fetchone()

            # Append vertaical summary
            if summary_row_vertical:
                row_idx = sheet.max_row + 1
                for col_idx, value in enumerate(summary_row_vertical, start=2):
                    sheet.cell(row=row_idx, column=col_idx, value=value)

            cursor.close()
            conn.close()
        except Exception as e:
            logger.error(f"Error fetching data for month {month}: {e}")
            return False
        
        # Second Block
        start_col_second = 12

        try:
            conn = get_connection()
            cursor = conn.cursor()
            logger.info(f"Fetching data for for month {month} from database...")
            
            cursor.execute("""
                SELECT 
                    FINAL_STATUS,
                    ISNULL(CAST(SUM([P20_DENOM]) * 1.0 / NULLIF(SUM(SUM([P20_DENOM])) OVER (), 0) AS DECIMAL(10,6)), 0) AS [P20_DENOM],
                    CAST(SUM([P50_DENOM]) * 1.0 / NULLIF(SUM(SUM([P50_DENOM])) OVER (), 0) AS DECIMAL(10,6)) AS [P50_DENOM],
                    CAST(SUM([P100_DENOM]) * 1.0 / NULLIF(SUM(SUM([P100_DENOM])) OVER (), 0) AS DECIMAL(10,6)) AS [P100_DENOM],
                    CAST(SUM([P200_DENOM]) * 1.0 / NULLIF(SUM(SUM([P200_DENOM])) OVER (), 0) AS DECIMAL(10,6)) AS [P200_DENOM],
                    CAST(SUM([P500_DENOM]) * 1.0 / NULLIF(SUM(SUM([P500_DENOM])) OVER (), 0) AS DECIMAL(10,6)) AS [P500_DENOM],
                    CAST(SUM([P1000_DENOM]) * 1.0 / NULLIF(SUM(SUM([P1000_DENOM])) OVER (), 0) AS DECIMAL(10,6)) AS [P1000_DENOM]            
                FROM [dbo].[txn_analysis_total_per_cash_bill_cbi_vol]
                WHERE TRANSACTION_DATE BETWEEN ? AND ?
                AND [MONTH] = ?
                GROUP BY FINAL_STATUS
                ORDER BY FINAL_STATUS;
            """, min_date_month, max_date_month, month)
            rows = cursor.fetchall()

            summary_query_vertical = """
                WITH totals AS (
                    SELECT
                        SUM([P20_DENOM]) AS [SUM_P20_DENOM],
                        SUM([P50_DENOM]) AS [SUM_P50_DENOM],
                        SUM([P100_DENOM]) AS [SUM_P100_DENOM],
                        SUM([P200_DENOM]) AS [SUM_P200_DENOM],
                        SUM([P500_DENOM]) AS [SUM_P500_DENOM],
                        SUM([P1000_DENOM]) AS [SUM_P1000_DENOM] 
                    FROM [dbo].[txn_analysis_total_per_cash_bill_cbi_vol]
                    WHERE TRANSACTION_DATE BETWEEN ? AND ?
                    AND [MONTH] = ?
                )
                SELECT
                    '',
                    ISNULL(CAST([SUM_P20_DENOM]  * 1.0 / NULLIF([SUM_P20_DENOM], 0)   AS DECIMAL(10,6)), 0) AS [P20_DENOM],
                    CAST([SUM_P50_DENOM] * 1.0 / NULLIF([SUM_P50_DENOM], 0)  AS DECIMAL(10,6)) AS [P50_DENOM],
                    CAST([SUM_P100_DENOM] * 1.0 / NULLIF([SUM_P100_DENOM], 0) AS DECIMAL(10,6)) AS [P100_DENOM],
                    CAST([SUM_P200_DENOM] * 1.0 / NULLIF([SUM_P200_DENOM], 0) AS DECIMAL(10,6)) AS [P200_DENOM],
                    CAST([SUM_P500_DENOM] * 1.0 / NULLIF([SUM_P500_DENOM], 0) AS DECIMAL(10,6)) AS [P500_DENOM],
                    CAST([SUM_P1000_DENOM] * 1.0 / NULLIF([SUM_P1000_DENOM], 0) AS DECIMAL(10,6)) AS [P1000_DENOM]
                FROM totals;
            """

            # Add headers 
            columns = [col[0] for col in cursor.description]

            # Style header row
            header_row_idx = sheet.max_row - len(rows) - 1
            for col_idx, header in enumerate(columns, start=start_col_second):
                cell = sheet.cell(row=header_row_idx, column=col_idx, value=header)
                if col_idx > 11  and col_idx < 20:
                    cell.font = header_font
                    cell.fill = header_fill
                    cell.border = border
                    cell.alignment = Alignment(horizontal="center")

            for i, row in enumerate(rows, start=1):
                row_idx = header_row_idx + i
                for col_idx, value in enumerate(tuple(row), start=start_col_second):
                    cell = sheet.cell(row=row_idx, column=col_idx, value=value)
                    if col_idx >= start_col_second and col_idx < start_col_second + 7:
                        cell.alignment = Alignment(horizontal="center")
                        cell.border = border
                        
                    if 13 <= col_idx <= 18:
                        cell.number_format = "0.00%"

            # Execute Vertical Summary
            cursor.execute(summary_query_vertical, min_date_month, max_date_month, month)
            summary_row_vertical = cursor.fetchone()
            
            if summary_row_vertical:
                row_idx = sheet.max_row
                for col_idx, value in enumerate(summary_row_vertical, start=start_col_second):
                    cell = sheet.cell(row=row_idx, column=col_idx, value=value)
                    cell.number_format = '0.00%'

            cursor.close()
            conn.close()
        except Exception as e:
            logger.error(f"Error fetching data for month {month}: {e}")
            return False
        
        #Third Block
        start_col_third = 20

        try:
            conn = get_connection()
            cursor = conn.cursor()
            logger.info(f"Fetching data for month {month} from database...")
            cursor.execute("""
                SELECT
                    FINAL_STATUS,
                    CAST(SUM([P20_DENOM])          * 1.0 / NULLIF((SUM([P20_DENOM]) + SUM([P50_DENOM]) + SUM([P100_DENOM]) + SUM([P200_DENOM]) + SUM([P500_DENOM]) + SUM([P1000_DENOM])), 0) AS DECIMAL(10,6)) AS [P20_DENOM],
                    CAST(SUM([P50_DENOM])     * 1.0 / NULLIF((SUM([P20_DENOM]) + SUM([P50_DENOM]) + SUM([P100_DENOM]) + SUM([P200_DENOM]) + SUM([P500_DENOM]) + SUM([P1000_DENOM])), 0) AS DECIMAL(10,6)) AS [P50_DENOM],
                    CAST(SUM([P100_DENOM])    * 1.0 / NULLIF((SUM([P20_DENOM]) + SUM([P50_DENOM]) + SUM([P100_DENOM]) + SUM([P200_DENOM]) + SUM([P500_DENOM]) + SUM([P1000_DENOM])), 0) AS DECIMAL(10,6)) AS [P100_DENOM],
                    CAST(SUM([P200_DENOM]) * 1.0 / NULLIF((SUM([P20_DENOM]) + SUM([P50_DENOM]) + SUM([P100_DENOM]) + SUM([P200_DENOM]) + SUM([P500_DENOM]) + SUM([P1000_DENOM])), 0) AS DECIMAL(10,6)) AS [P200_DENOM],
                    CAST(SUM([P500_DENOM]) * 1.0 / NULLIF((SUM([P20_DENOM]) + SUM([P50_DENOM]) + SUM([P100_DENOM]) + SUM([P200_DENOM]) + SUM([P500_DENOM]) + SUM([P1000_DENOM])), 0) AS DECIMAL(10,6)) AS [P500_DENOM],
                    CAST(SUM([P1000_DENOM]) * 1.0 / NULLIF((SUM([P20_DENOM]) + SUM([P50_DENOM]) + SUM([P100_DENOM]) + SUM([P200_DENOM]) + SUM([P500_DENOM]) + SUM([P1000_DENOM])), 0) AS DECIMAL(10,6)) AS [P1000_DENOM],
                    CAST(1.0 AS DECIMAL(10,6))
                FROM [dbo].[txn_analysis_total_per_cash_bill_cbi_vol]
                WHERE TRANSACTION_DATE BETWEEN ? AND ?
                AND [MONTH] = ?
                GROUP BY FINAL_STATUS
                ORDER BY FINAL_STATUS
            """, min_date_month, max_date_month, month)
            rows = cursor.fetchall()

            # Add headers 
            columns = [col[0] for col in cursor.description]

            # Style header row
            header_row_idx = sheet.max_row - len(rows) - 1
            for col_idx, header in enumerate(columns, start=start_col_third):
                cell = sheet.cell(row=header_row_idx, column=col_idx, value=header)
                if col_idx > 15 and col_idx < 27:
                    cell.font = header_font
                    cell.fill = header_fill
                    cell.border = border
                    cell.alignment = Alignment(horizontal="center")

            for i, row in enumerate(rows, start=1):
                row_idx = header_row_idx + i
                for col_idx, value in enumerate(tuple(row), start=start_col_third):
                    cell = sheet.cell(row=row_idx, column=col_idx, value=value)
                    if col_idx >= start_col_third and col_idx < start_col_third + 7:
                        cell.alignment = Alignment(horizontal="center")
                        cell.border = border

                    if 16 <= col_idx <= 27:
                        cell.number_format = "0.00%"

            cursor.close()
            conn.close()
        except Exception as e:
            logger.error(f"Error fetching data for month {month}: {e}")
            return False

        
        # Add a blank row after each day's block
        if rows:
            sheet.append([])


        
    logger.info("TOTAL_PER CASH BILL - CBI (MONTHLY)  data generation complete.")
    return True

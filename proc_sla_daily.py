from openpyxl.worksheet.worksheet import Worksheet
from conn import get_connection
import logging
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment

# Define styles
title_font = Font(bold=True, color="008000", size=11)  # Green bold title
header_font = Font(bold=True, color="FFFFFF")          # White bold header
header_fill = PatternFill(start_color="156082", end_color="156082", fill_type="solid")  # Light green fill
border = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin")
)

def generate_sla_daily(
    sheet: Worksheet,
    end_date: str, 
    logger: logging.Logger
) -> bool:
    """Generate SLA (DAILY) data into the given Excel sheet."""
    logger.info("Generating SLA (DAILY) data...")

    # MIN AND MAX DATES
    min_date = end_date[:8] + "01"
    max_date = end_date

    # Add title
    title_cell = sheet.cell(row=1, column=2, value="TRANSACTION COUNT PER SLA- IN SECONDS AND FINAL STATUS")
    title_cell_2 = sheet.cell(row=1, column=9, value="SLA- (UPDATED DATE - CREATED DATE)")
    
    title_cell.font = title_font
    title_cell_2.font = title_font

    for day in range(int(min_date[8:]), int(max_date[8:]) + 1):
        current_date = f"{min_date[:8]}{day:02d}"
        logger.info(f"Processing data for date: {current_date}")

        # First Block
        try:
            conn = get_connection()
            cursor = conn.cursor()
            logger.info(f"Fetching data for {current_date} from database for first block...")
            cursor.execute("""
                SELECT 
                [WEEK_NUM],
                [FINAL_STATUS],
                [SLA-<0],
                [SLA-=0],
                [SLA-1TO30],
                [SLA-31TO60],
                [SLA-61TO90],
                [SLA-91TO120],
                [SLA-121TO150],
                [SLA->150],                                                                                 
                [SLA-<0] + [SLA-=0] + [SLA-1TO30] + [SLA-31TO60] + [SLA-61TO90] + [SLA-91TO120] + [SLA-121TO150] + [SLA->150]
                FROM [dbo].[txn_analysis_sla]
                WHERE TRANSACTION_DATE = ?
                ORDER BY WEEK_NUM, FINAL_STATUS
            """, current_date)
            rows = cursor.fetchall()

            summary_query_vertical = """
                SELECT 
                '',
                '',
                SUM([SLA-<0]) AS [SLA-<0],
                SUM([SLA-=0]) AS [SLA-=0],
                SUM([SLA-1TO30]) AS [SLA-1TO30],
                SUM([SLA-31TO60]) AS [SLA-31TO60],
                SUM([SLA-61TO90]) AS [SLA-61TO90],
                SUM([SLA-91TO120]) AS [SLA-91TO120],
                SUM([SLA-121TO150]) AS [SLA-121TO150],
                SUM([SLA->150]) AS [SLA->150]
                FROM [dbo].[txn_analysis_sla]
                WHERE TRANSACTION_DATE = ?
            """

            # Add headers 
            columns = [col[0] for col in cursor.description]
            columns.insert(0, current_date)
            sheet.append(columns)

            # Style header row
            header_row_idx = sheet.max_row
            for col_idx, _ in enumerate(columns, start=1):
                cell = sheet.cell(row=header_row_idx, column=col_idx)
                if col_idx > 1 and col_idx < 12:
                    cell.font = header_font
                    cell.fill = header_fill
                    cell.border = border
                    cell.alignment = Alignment(horizontal="center")

            # Append each row
            for row in rows:
                sheet.append(('',) + tuple(row))
                row_idx = sheet.max_row
                for col_idx in range(1, len(columns) + 1):
                    cell = sheet.cell(row=row_idx, column=col_idx)
                    if col_idx > 1 and col_idx < 12:
                        cell.alignment = Alignment(horizontal="center")
                        cell.border = border

            # Execute Vertical Summary
            cursor.execute(summary_query_vertical, current_date)
            summary_row_vertical = cursor.fetchone()

            # Append vertaical summary
            if summary_row_vertical:
                row_idx = sheet.max_row + 1
                for col_idx, value in enumerate(summary_row_vertical, start=2):
                    sheet.cell(row=row_idx, column=col_idx, value=value)

            cursor.close()
            conn.close()
        except Exception as e:
            logger.error(f"Error fetching data for first block {current_date}: {e}")
            return False
        
        #Second Block
        start_col_second = 14

        try:
            conn = get_connection()
            cursor = conn.cursor()
            logger.info(f"Fetching data for second block {current_date} from database...")
            
            cursor.execute("""
                SELECT 
                    FINAL_STATUS,
                    CAST([SLA-<0] * 1.0 / NULLIF(SUM([SLA-<0]) OVER (), 0) AS DECIMAL(10,6)) AS [SLA-<0],
                    CAST([SLA-=0] * 1.0 / NULLIF(SUM([SLA-=0]) OVER (), 0) AS DECIMAL(10,6)) AS [SLA-=0],
                    CAST([SLA-1TO30] * 1.0 / NULLIF(SUM([SLA-1TO30]) OVER (), 0) AS DECIMAL(10,6)) AS [SLA-1TO30],
                    CAST([SLA-31TO60] * 1.0 / NULLIF(SUM([SLA-31TO60]) OVER (), 0) AS DECIMAL(10,6)) AS [SLA-31TO60],
                    CAST([SLA-61TO90] * 1.0 / NULLIF(SUM([SLA-61TO90]) OVER (), 0) AS DECIMAL(10,6)) AS [SLA-61TO90],
                    CAST([SLA-91TO120] * 1.0 / NULLIF(SUM([SLA-91TO120]) OVER (), 0) AS DECIMAL(10,6)) AS [SLA-91TO120],
                    CAST([SLA-121TO150] * 1.0 / NULLIF(SUM([SLA-121TO150]) OVER (), 0) AS DECIMAL(10,6)) AS [SLA-121TO150],
                    CAST([SLA->150] * 1.0 / NULLIF(SUM([SLA->150]) OVER (), 0) AS DECIMAL(10,6)) AS [SLA->150]         
                FROM [dbo].[txn_analysis_sla]
                WHERE TRANSACTION_DATE = ?
                ORDER BY WEEK_NUM, FINAL_STATUS;
            """, current_date)
            rows = cursor.fetchall()

            summary_query_vertical = """
                WITH totals AS (
                    SELECT
                        SUM([SLA-<0])  AS [SUM_SLA-<0],
                        SUM([SLA-=0]) AS [SUM_SLA-=0],
                        SUM([SLA-1TO30]) AS [SUM_SLA-1TO30],
                        SUM([SLA-31TO60]) AS [SUM_SLA-31TO60],
                        SUM([SLA-61TO90]) AS [SUM_SLA-61TO90],
                        SUM([SLA-91TO120]) AS [SUM_SLA-91TO120],
                        SUM([SLA-121TO150]) AS [SUM_SLA-121TO150],
                        SUM([SLA->150]) AS [SLA->150]
                    FROM [dbo].[txn_analysis_sla]
                    WHERE TRANSACTION_DATE = ?
                )
                SELECT
                    '',
                    CAST([SUM_SLA-<0]  * 1.0 / NULLIF([SUM_SLA-<0], 0)   AS DECIMAL(10,6)) AS [SUM_SLA-<0],
                    CAST([SUM_SLA-=0] * 1.0 / NULLIF([SUM_SLA-=0], 0)  AS DECIMAL(10,6)) AS [SUM_SLA-=0],
                    CAST([SUM_SLA-1TO30] * 1.0 / NULLIF([SUM_SLA-1TO30], 0) AS DECIMAL(10,6)) AS [SUM_SLA-1TO30],
                    CAST([SUM_SLA-31TO60] * 1.0 / NULLIF([SUM_SLA-31TO60], 0) AS DECIMAL(10,6)) AS [SUM_SLA-31TO60],
                    CAST([SUM_SLA-61TO90] * 1.0 / NULLIF([SUM_SLA-61TO90], 0) AS DECIMAL(10,6)) AS [SUM_SLA-61TO90],
                    CAST([SUM_SLA-91TO120] * 1.0 / NULLIF([SUM_SLA-91TO120], 0) AS DECIMAL(10,6)) AS [SUM_SLA-91TO120],
                    CAST([SUM_SLA-121TO150] * 1.0 / NULLIF([SUM_SLA-121TO150], 0) AS DECIMAL(10,6)) AS [SUM_SLA-121TO150],
                    CAST([SLA->150] * 1.0 / NULLIF([SLA->150], 0) AS DECIMAL(10,6)) AS [SLA->150]
                FROM totals;
            """

            # Add headers 
            columns = [col[0] for col in cursor.description]

            # Style header row
            header_row_idx = sheet.max_row - len(rows) - 1
            for col_idx, header in enumerate(columns, start=start_col_second):
                cell = sheet.cell(row=header_row_idx, column=col_idx, value=header)
                if col_idx > 13  and col_idx < 23:
                    cell.font = header_font
                    cell.fill = header_fill
                    cell.border = border
                    cell.alignment = Alignment(horizontal="center")

            for i, row in enumerate(rows, start=1):
                row_idx = header_row_idx + i
                for col_idx, value in enumerate(tuple(row), start=start_col_second):
                    cell = sheet.cell(row=row_idx, column=col_idx, value=value)
                    if col_idx >= start_col_second and col_idx < start_col_second + 9:
                        cell.alignment = Alignment(horizontal="center")
                        cell.border = border
                        
                    if 15 <= col_idx <= 22:
                        cell.number_format = "0.00%"

            # Execute Vertical Summary
            cursor.execute(summary_query_vertical, current_date)
            summary_row_vertical = cursor.fetchone()
            
            if summary_row_vertical:
                row_idx = sheet.max_row
                for col_idx, value in enumerate(summary_row_vertical, start=start_col_second):
                    cell = sheet.cell(row=row_idx, column=col_idx, value=value)
                    cell.number_format = '0.00%'

            cursor.close()
            conn.close()
        except Exception as e:
            logger.error(f"Error fetching data for second block {current_date}: {e}")
            return False
        
        #Third Block
        start_col_third = 24

        try:
            conn = get_connection()
            cursor = conn.cursor()
            logger.info(f"Fetching data for third block {current_date} from database...")
            cursor.execute("""
                SELECT
                    FINAL_STATUS,
                    CAST([SLA-<0]   * 1.0 / NULLIF(([SLA-<0]     + [SLA-=0] + [SLA-1TO30] + [SLA-31TO60] + [SLA-61TO90] + [SLA-91TO120] + [SLA-121TO150] + [SLA->150]), 0) AS DECIMAL(10,6)) AS [SLA-<0],
                    CAST([SLA-=0]  * 1.0 / NULLIF(([SLA-<0]      + [SLA-=0] + [SLA-1TO30] + [SLA-31TO60] + [SLA-61TO90] + [SLA-91TO120] + [SLA-121TO150] + [SLA->150]), 0) AS DECIMAL(10,6)) AS [SLA-=0],
                    CAST([SLA-1TO30] * 1.0 / NULLIF(([SLA-<0]    + [SLA-=0] + [SLA-1TO30] + [SLA-31TO60] + [SLA-61TO90] + [SLA-91TO120] + [SLA-121TO150] + [SLA->150]), 0) AS DECIMAL(10,6)) AS [SLA-1TO30],
                    CAST([SLA-31TO60] * 1.0 / NULLIF(([SLA-<0]   + [SLA-=0] + [SLA-1TO30] + [SLA-31TO60] + [SLA-61TO90] + [SLA-91TO120] + [SLA-121TO150] + [SLA->150]), 0) AS DECIMAL(10,6)) AS [SLA-31TO60],
                    CAST([SLA-61TO90] * 1.0 / NULLIF(([SLA-<0]   + [SLA-=0] + [SLA-1TO30] + [SLA-31TO60] + [SLA-61TO90] + [SLA-91TO120] + [SLA-121TO150] + [SLA->150]), 0) AS DECIMAL(10,6)) AS [SLA-61TO90],
                    CAST([SLA-91TO120] * 1.0 / NULLIF(([SLA-<0]  + [SLA-=0] + [SLA-1TO30] + [SLA-31TO60] + [SLA-61TO90] + [SLA-91TO120] + [SLA-121TO150] + [SLA->150]), 0) AS DECIMAL(10,6)) AS [SLA-91TO120],
                    CAST([SLA-121TO150] * 1.0 / NULLIF(([SLA-<0] + [SLA-=0] + [SLA-1TO30] + [SLA-31TO60] + [SLA-61TO90] + [SLA-91TO120] + [SLA-121TO150] + [SLA->150]), 0) AS DECIMAL(10,6)) AS [SLA-121TO150],
                    CAST([SLA->150] * 1.0 / NULLIF(([SLA-<0]     + [SLA-=0] + [SLA-1TO30] + [SLA-31TO60] + [SLA-61TO90] + [SLA-91TO120] + [SLA-121TO150] + [SLA->150]), 0) AS DECIMAL(10,6)) AS [SLA->150],
                    CAST(1.0 AS DECIMAL(10,6))
                FROM [dbo].[txn_analysis_sla]
                WHERE TRANSACTION_DATE = ?
                ORDER BY WEEK_NUM, FINAL_STATUS;
            """, current_date)
            rows = cursor.fetchall()

            # Add headers 
            columns = [col[0] for col in cursor.description]

            # Style header row
            header_row_idx = sheet.max_row - len(rows) - 1
            for col_idx, header in enumerate(columns, start=start_col_third):
                cell = sheet.cell(row=header_row_idx, column=col_idx, value=header)
                if col_idx > 23 and col_idx < 33:
                    cell.font = header_font
                    cell.fill = header_fill
                    cell.border = border
                    cell.alignment = Alignment(horizontal="center")

            for i, row in enumerate(rows, start=1):
                row_idx = header_row_idx + i
                for col_idx, value in enumerate(tuple(row), start=start_col_third):
                    cell = sheet.cell(row=row_idx, column=col_idx, value=value)
                    if col_idx >= start_col_third and col_idx < start_col_third + 9:
                        cell.alignment = Alignment(horizontal="center")
                        cell.border = border

                    if 24 <= col_idx <= 33:
                        cell.number_format = "0.00%"

            cursor.close()
            conn.close()
        except Exception as e:
            logger.error(f"Error fetching data for third block {current_date}: {e}")
            return False

        
        # Add a blank row after each day's block
        if rows:
            sheet.append([])
        
    logger.info("SLA (DAILY) data generation complete.")
    return True
